#!/usr/bin/env python3
"""Genera checklist-revision-codigo-sprintops.xlsx desde el CSV."""
import csv
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Instala: pip install openpyxl")
    raise

ROOT = Path(__file__).resolve().parent
CSV_PATH = ROOT / "checklist-revision-codigo-sprintops.csv"
XLSX_PATH = ROOT / "checklist-revision-codigo-sprintops.xlsx"

FILL = {
    "Cumple": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "No cumple": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "Parcial": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
}
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
BORDER = Border(*(Side(style="thin"),) * 4)
WRAP = Alignment(wrap_text=True, vertical="top")

wb = Workbook()
ws = wb.active
ws.title = "Checklist revision"

with CSV_PATH.open(encoding="utf-8") as f:
    for row in csv.reader(f, delimiter=";"):
        ws.append(row)

# Estilos fila 7 = encabezados tabla
for c in range(1, 6):
    cell = ws.cell(7, c)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal="center", wrap_text=True)
    cell.border = BORDER

for r in range(8, ws.max_row + 1):
    res = ws.cell(r, 4).value
    fill = FILL.get(res)
    for c in range(1, 6):
        cell = ws.cell(r, c)
        cell.alignment = WRAP
        cell.border = BORDER
        if fill and c >= 4:
            cell.fill = fill

for i, w in enumerate([5, 22, 50, 12, 55], 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A8"
wb.save(XLSX_PATH)
print(XLSX_PATH)
